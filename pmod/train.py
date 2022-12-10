import argparse
import codecs
import datetime
import os
import shutil
from typing import Dict, Union
import yaml
from glob import glob
import optuna
from tqdm import tqdm
import numpy as np
import openpyxl as xl
import torch.nn as nn
from torch import Tensor
from torch.cuda.amp import autocast, GradScaler
from torch.utils import tensorboard
from torchvision.utils import make_grid
from pointsmap import depth2colormap
from .model.constant import *
from .model.model import PMOD
from .model.metric import *
from .model.utils import *
from .dataloader import *
from .evaluate import main as evaluate


def parse_args() -> dict:
    parser = argparse.ArgumentParser()

    parser_train = parser.add_argument_group('Training')
    parser_train.add_argument(
        '-t', f'--{arg_hyphen(ARG_TAG)}',
        type=str, required=True,
        help='Training Tag.'
    )
    parser_train.add_argument(
        '-tdc', f'--{arg_hyphen(ARG_TRAIN_DL_CONFIG)}',
        type=str, metavar='PATH', required=True,
        help='PATH of JSON file of dataloader config for training.'
    )
    parser_train.add_argument(
        '-vdc', f'--{arg_hyphen(ARG_VAL_DL_CONFIG)}',
        type=str, metavar='PATH', default=None,
        help=f'PATH of JSON file of dataloader config for validation. If not specified, the same file as "--{arg_hyphen(ARG_TRAIN_DL_CONFIG)}" will be used.'
    )
    parser_train.add_argument(
        '-bs', f'--{arg_hyphen(ARG_BLOCK_SIZE)}',
        type=int, default=0,
        help='Block size of dataset.'
    )
    parser_train.add_argument(
        '-td', f'--{arg_hyphen(ARG_TRAIN_DATA)}',
        type=str, metavar='PATH', nargs='+', required=True,
        help='PATH of training HDF5 datasets.'
    )
    parser_train.add_argument(
        '-vd', f'--{arg_hyphen(ARG_VAL_DATA)}',
        type=str, metavar='PATH', nargs='*', default=[],
        help=f'PATH of validation HDF5 datasets. If not specified, the same files as "--{arg_hyphen(ARG_TRAIN_DATA)}" will be used.'
    )
    parser_train.add_argument(
        f'--{arg_hyphen(ARG_EPOCHS)}',
        type=int, default=200,
        help='Epochs'
    )
    parser_train.add_argument(
        f'--{arg_hyphen(ARG_EPOCH_START_COUNT)}',
        type=int, default=1,
        help='The starting epoch count'
    )
    parser_train.add_argument(
        f'--{arg_hyphen(ARG_STEPS_PER_EPOCH)}',
        type=int, default=10000,
        help='Number of steps per epoch. If it is greater than the total number of datasets, then the total number of datasets is used.'
    )
    parser_train.add_argument(
        '-ppa', f'--{arg_hyphen(ARG_PROJECTED_POSITION_AUGMENTATION)}',
        action='store_false', help='Unuse Projected Positiion Augmentation'
    )
    parser_train.add_argument(
        f'--{arg_hyphen(ARG_TR_ERROR_RANGE)}',
        type=float, nargs=3, default=[0.6, 1.3, 0.7],
        help='Translation Error Range [m].'
    )
    parser_train.add_argument(
        f'--{arg_hyphen(ARG_ROT_ERROR_RANGE)}',
        type=float, default=3.0,
        help='Rotation Error Range [deg].'
    )
    parser_train.add_argument(
        '-ae', f'--{arg_hyphen(ARG_AUTO_EVALUATION)}',
        action='store_true',
        help='Auto Evaluation.'
    )
    parser_train.add_argument(
        '-edc', f'--{arg_hyphen(ARG_EVAL_DL_CONFIG)}',
        type=str, metavar='PATH',
        help='PATH of JSON file of dataloader config.'
    )
    parser_train.add_argument(
        '-ed', f'--{arg_hyphen(ARG_EVAL_DATA)}',
        type=str, metavar='PATH', nargs='*', default=[],
        help='PATH of evaluation HDF5 datasets.'
    )
    parser_train.add_argument(
        f'--{arg_hyphen(ARG_NOMAP)}',
        action='store_true',
        help='No map input.'
    )
    parser_train.add_argument(
        f'--{arg_hyphen(ARG_THRESHOLDS)}',
        type=float, nargs='+', default=DEFAULT_THRESHOLDS,
        help='Thresholds of depth.'
    )
    parser_train.add_argument(
        f'--{arg_hyphen(ARG_SEED)}',
        type=int, default=0,
        help='Random seed.'
    )

    parser_net = parser.add_argument_group('Network')
    parser_net.add_argument(
        '-b', f'--{arg_hyphen(ARG_BATCH_SIZE)}',
        type=int, default=4,
        help='Batch Size'
    )
    parser_net.add_argument(
        f'--{arg_hyphen(ARG_RESUME)}',
        type=str, metavar='PATH', default=None,
        help='PATH of checkpoint(.pth).'
    )
    parser_net.add_argument(
        f'-amp', f'--{arg_hyphen(ARG_AMP)}',
        action='store_true',
        help='Use AMP.'
    )

    parser_optim = parser.add_argument_group('Optimizer')
    parser_optim.add_argument(
        f'--{arg_hyphen(ARG_CLIP_MAX_NORM)}',
        type=float, default=1.0,
        help='max_norm for clip_grad_norm.'
    )
    parser_optim.add_argument(
        '-op', f'--{arg_hyphen(ARG_OPTIM_PARAMS)}',
        type=str, metavar='PATH', default='./config/optim-params-default.yaml',
        help='PATH of YAML file of optimizer params.'
    )
    parser_optim.add_argument(
        '-o', f'--{arg_hyphen(ARG_OPTIMIZER)}',
        type=str, default=OPTIM_TYPE_ADABELIEF, choices=OPTIM_TYPES,
        help='Optimizer'
    )
    parser_optim.add_argument(
        '-lp', f'--{arg_hyphen(ARG_LR_POLICY)}',
        type=str, default=LR_POLICY_PLATEAU, choices=LR_POLICIES,
        help='Learning rate policy.'
    )

    parser_loss = parser.add_argument_group('Loss')
    parser_loss.add_argument(
        f'--{arg_hyphen(ARG_L1)}',
        type=float, default=5.0,
        help='Weight of L1 loss.'
    )
    parser_loss.add_argument(
        f'--{arg_hyphen(ARG_SEG_CE)}',
        type=float, default=0.5,
        help='Weight of Segmentation CrossEntropy Loss.'
    )
    parser_loss.add_argument(
        f'--{arg_hyphen(ARG_SEG_CE_AUX1)}',
        type=float, default=0.3,
        help='Weight of Segmentation Aux1 CrosEntropy Loss.'
    )
    parser_loss.add_argument(
        f'--{arg_hyphen(ARG_SEG_CE_AUX2)}',
        type=float, default=0.25,
        help='Weight of Segmentation Aux2 CrosEntropy Loss.'
    )

    parser_debug = parser.add_argument_group('Debug')
    parser_debug.add_argument(
        f'--{arg_hyphen(ARG_DETECT_ANOMALY)}',
        action='store_true',
        help='AnomalyMode'
    )

    args: dict = vars(parser.parse_args())

    if os.path.isfile(args[ARG_OPTIM_PARAMS]) is False:
        raise FileNotFoundError(f'"{args[ARG_OPTIM_PARAMS]}"')
    with open(args[ARG_OPTIM_PARAMS]) as f:
        optim_params: dict = yaml.safe_load(f)
    args[ARG_OPTIM_PARAMS] = optim_params

    return args


def main(args: Dict[str, Union[int, float, str, dict]], workdir: str, trial: optuna.Trial = None) -> Dict[str, float]:
    if os.path.isdir(workdir) is False:
        raise NotADirectoryError(f'{workdir}')

    ###############
    # Random Seed #
    ###############
    np.random.seed(seed=args[ARG_SEED])
    torch.random.manual_seed(seed=args[ARG_SEED])

    ##################
    # Device Setting #
    ##################
    if torch.cuda.is_available():
        device: torch.device = torch.device('cuda')
        torch.backends.cudnn.benchmark = True
        torch.cuda.manual_seed_all(seed=args[ARG_SEED])
        torch.backends.cudnn.deterministic = True
        print(f'{"Device":11s}: {torch.cuda.get_device_name(device)}')
    else:
        device: torch.device = torch.device('cpu')
        print(f'{"Device":11s}: CPU')

    ######################
    # DataLoader Setting #
    ######################
    if args[ARG_VAL_DL_CONFIG] is None:
        args[ARG_VAL_DL_CONFIG] = args[ARG_TRAIN_DL_CONFIG]
    if len(args[ARG_VAL_DATA]) == 0:
        args[ARG_VAL_DATA] = args[ARG_TRAIN_DATA]
    if args.get(ARG_EVAL_DL_CONFIG) is None:
        args[ARG_EVAL_DL_CONFIG] = args[ARG_VAL_DL_CONFIG]
    if len(args[ARG_EVAL_DATA]) == 0:
        args[ARG_EVAL_DATA] = args[ARG_VAL_DATA]
    if args[ARG_BLOCK_SIZE] < 3:
        args[ARG_BLOCK_SIZE] = 0

    with redirect_stdout(open(os.devnull, 'w')):
        train_dataloader: DataLoader = init_train_dataloader(args)
        val_dataloader: DataLoader = init_val_dataloader(args)
    train_dataset: PMOD_Train_Dataset = train_dataloader.dataset
    label_tag: str = train_dataset.minibatch[DATASET_LABEL][CONFIG_TAG_LABELTAG]

    steps_per_epoch: int = args[ARG_STEPS_PER_EPOCH]
    train_data_len: int = len(train_dataloader)
    if train_data_len < steps_per_epoch:
        steps_per_epoch = train_data_len
        args[ARG_STEPS_PER_EPOCH] = steps_per_epoch

    ###################
    # Network Setting #
    ###################
    if args[ARG_RESUME] is None:
        pmodnet: nn.Module = PMOD(
            camera_shape=args[ARG_CAMERA_SHAPE],
            num_classes=args[ARG_NUM_CLASSES]
        ).to(device)
    else:
        if os.path.isfile(args[ARG_RESUME]) is False:
            raise FileNotFoundError(f'{args[ARG_RESUME]}')
        pmodnet: nn.Module = torch.load(args[ARG_RESUME])
        pmodnet.to(device)

    ################
    # Count Params #
    ################
    params = 0
    for p in pmodnet.parameters():
        if p.requires_grad is True:
            params += p.numel()
    args[ARG_PARAMS] = params

    #####################
    # Optimizer Setting #
    #####################
    optim_params: Dict[str, dict] = args[ARG_OPTIM_PARAMS]
    # Encoder
    optimizer_enc, scheduler_enc = init_optimizer(
        net=pmodnet.encoder, optimizer_type=args[ARG_OPTIMIZER], scheduler_type=args[ARG_LR_POLICY],
        config=optim_params[CONFIG_ENCODER], steps_per_epoch=steps_per_epoch
    )
    scaler_enc = GradScaler(enabled=args[ARG_AMP])
    # Decoder - Depth
    optimizer_dec_depth, scheduler_dec_depth = init_optimizer(
        net=pmodnet.decoder_depth, optimizer_type=args[ARG_OPTIMIZER], scheduler_type=args[ARG_LR_POLICY],
        config=optim_params[f'{CONFIG_DECODER}_{CONFIG_DEPTH}'], steps_per_epoch=steps_per_epoch
    )
    scaler_dec_depth = GradScaler(enabled=args[ARG_AMP])
    # Decoder - Segmentation
    optimizer_dec_seg, scheduler_dec_seg = init_optimizer(
        net=pmodnet.decoder_seg, optimizer_type=args[ARG_OPTIMIZER], scheduler_type=args[ARG_LR_POLICY],
        config=optim_params[f'{CONFIG_DECODER}_{CONFIG_SEGMENTATION}'], steps_per_epoch=steps_per_epoch
    )
    scaler_dec_seg = GradScaler(enabled=args[ARG_AMP])

    ################
    # Loss Setting #
    ################
    lossL1: LOSS_WEIGHT = LOSS_WEIGHT(
        LossL1() if args[ARG_L1] != 0.0 else LossNULL(), args[ARG_L1]).to(device)
    lossCE: LOSS_WEIGHT = LOSS_WEIGHT(
        nn.CrossEntropyLoss(), args[ARG_SEG_CE]).to(device)
    lossCE_aux1: LOSS_WEIGHT = LOSS_WEIGHT(
        nn.CrossEntropyLoss(), args[ARG_SEG_CE_AUX1]).to(device)
    lossCE_aux2: LOSS_WEIGHT = LOSS_WEIGHT(
        nn.CrossEntropyLoss(), args[ARG_SEG_CE_AUX2]).to(device)

    ##################
    # Metric Setting #
    ##################
    metricSumIntersection: nn.Module = MetricSumIntersection().to(device)
    metricSumUnion: nn.Module = MetricSumUnion().to(device)
    metricSumAE: nn.Module = MetricSumAE().to(device)
    metricSumSE: nn.Module = MetricSumSE().to(device)
    metricSumAPE: nn.Module = MetricSumAPE().to(device)

    best_miou = METRIC_BEST(0.0, '', 0)
    best_mape = METRIC_BEST(np.inf, '', 0)

    eval_best_miou = METRIC_BEST(0.0, '', 0)
    eval_best_mape = METRIC_BEST(np.inf, '', 0)

    #################
    # Debug Setting #
    #################
    torch.autograd.set_detect_anomaly(args[ARG_DETECT_ANOMALY])

    ###############
    # Save Config #
    ###############
    if isinstance(trial, optuna.Trial):
        dt_start = trial.datetime_start
    else:
        dt_start = datetime.datetime.now()
    args[ARG_DATE] = dt_start.strftime('%Y%m%dT%H%M%S')
    train_name: str = f'{args[ARG_DATE]}-{args[ARG_TAG]}'
    checkpoint_dir: str = os.path.join(workdir, DIR_CHECKPOINTS, train_name)
    val_book_path: str = os.path.join(checkpoint_dir, 'validation.xlsx')

    print(f'{"CheckPoint":11s}: "{checkpoint_dir}"')

    #######################
    # TensorBoard Setting #
    #######################
    tb_log_dir: str = os.path.join(workdir, DIR_LOGS, train_name)
    tb_writer = None

    print(f'{"Summary":11s}: "{tb_log_dir}"')

    try:
        loss_dict: Dict[str, Dict[str, Tensor]] = {}
        #################
        # Training Loop #
        #################
        for epoch_itr in range(args[ARG_EPOCHS]):
            epoch_num: int = epoch_itr + 1
            ############
            # Training #
            ############
            pmodnet.train()

            for batch_itr in tqdm(range(steps_per_epoch), desc=f'Epoch {epoch_num:5d}: {"Train":14s}'):
                step_itr: int = epoch_itr * steps_per_epoch + batch_itr

                batch: Dict[str, Tensor] = next(iter(train_dataloader))
                in_camera: Tensor = batch[DATASET_CAMERA].to(
                    device, non_blocking=True)
                in_map: Tensor = batch[DATASET_MAP].to(
                    device, non_blocking=True)
                gt_label: Tensor = batch[DATASET_LABEL].to(
                    device, non_blocking=True)
                gt_depth: Tensor = batch[DATASET_DEPTH].to(
                    device, non_blocking=True)

                with autocast(enabled=args[ARG_AMP]):
                    pred: PMOD_OUTPUT = pmodnet(
                        camera=in_camera, map_depth=in_map)

                train_loss_dict: Dict[str, float] = {}

                optimizer_enc.zero_grad(set_to_none=True)
                optimizer_dec_depth.zero_grad(set_to_none=True)
                optimizer_dec_seg.zero_grad(set_to_none=True)

                with autocast(enabled=args[ARG_AMP]):
                    loss_seg_ce: Tensor = lossCE(pred.segmentation, gt_label)
                    train_loss_dict['CrossEntropy'] = loss_seg_ce.clone().detach()
                    loss_seg_aux1_ce: Tensor = lossCE_aux1(pred.aux1, gt_label)
                    train_loss_dict['Aux1'] = loss_seg_aux1_ce.clone().detach()
                    loss_seg_aux2_ce: Tensor = lossCE_aux2(pred.aux2, gt_label)
                    train_loss_dict['Aux2'] = loss_seg_aux2_ce.clone().detach()
                    loss_depth_l1: Tensor = lossL1(pred.depth, gt_depth)
                    train_loss_dict['L1'] = loss_depth_l1.clone().detach()

                    loss_g: Tensor = loss_seg_ce + loss_depth_l1 + \
                        loss_seg_aux1_ce + loss_seg_aux2_ce
                    train_loss_dict['All'] = loss_g.clone().detach()

                scaler_dec_seg.scale(loss_g)
                scaler_dec_depth.scale(loss_g)
                scaler_enc.scale(loss_g).backward()

                del loss_g, loss_seg_ce, loss_depth_l1, loss_seg_aux1_ce, loss_seg_aux2_ce

                nn.utils.clip_grad_norm_(
                    pmodnet.parameters(), max_norm=args[ARG_CLIP_MAX_NORM])

                scaler_enc.step(optimizer_enc)
                scaler_enc.update()
                scaler_dec_depth.step(optimizer_dec_depth)
                scaler_dec_depth.update()
                scaler_dec_seg.step(optimizer_dec_seg)
                scaler_dec_seg.update()

                loss_dict = {'Training': train_loss_dict}

                ######################
                # Update Tensorboard #
                ######################
                if (step_itr % 50 == 49):
                    if tb_writer is None:
                        os.makedirs(tb_log_dir, exist_ok=True)
                        tb_writer = tensorboard.SummaryWriter(
                            log_dir=tb_log_dir)
                    for h1, dicts in loss_dict.items():
                        for h2, value in dicts.items():
                            tb_writer.add_scalar(
                                f'{h1}/{h2}', value.item(), global_step=step_itr)
                    pred_seg: Tensor = torch.argmax(pred.segmentation, dim=1)
                    img_tensor: Tensor = make_grid([
                        torch.from_numpy(np.uint8(train_dataset.to_numpy(
                            batch[DATASET_CAMERA][0], DATASET_CAMERA)[..., ::-1] * 255).copy().transpose([2, 0, 1])),
                        torch.from_numpy(depth2colormap(train_dataset.to_numpy(
                            batch[DATASET_MAP][0], DATASET_MAP), 0.0, 1.0, invert=True)[..., ::-1].copy().transpose([2, 0, 1])),
                        torch.from_numpy(train_dataset.convert_semantic2d_to_rgb8(
                            pred_seg[0].cpu().detach().numpy(), label_tag).copy().transpose([2, 0, 1])),
                        torch.from_numpy(depth2colormap(train_dataset.to_numpy(
                            pred.depth[0], DATASET_DEPTH), 0.0, 1.0, invert=True)[..., ::-1].copy().transpose([2, 0, 1])),
                        torch.from_numpy(train_dataset.convert_semantic2d_to_rgb8(
                            batch[DATASET_LABEL][0], label_tag).copy().transpose([2, 0, 1])),
                        torch.from_numpy(depth2colormap(train_dataset.to_numpy(
                            batch[DATASET_DEPTH][0], DATASET_DEPTH), 0.0, 1.0, invert=True)[..., ::-1].copy().transpose([2, 0, 1])),
                    ], nrow=2)
                    tb_writer.add_image(
                        'Training/Images', img_tensor=img_tensor, global_step=step_itr)
                    tb_writer.flush()

            ##############
            # Validation #
            ##############
            pmodnet.eval()

            sum_intersection: float = 0.0
            sum_union: float = 0.0
            sum_ae: float = 0.0
            sum_se: float = 0.0
            sum_ape: float = 0.0
            count_steps: float = 0.0
            count_in_range: float = 0.0

            with torch.no_grad():
                for batch in tqdm(val_dataloader, desc=f'Epoch {epoch_num:5d}: {"Validation":14s}'):
                    in_camera: Tensor = batch[DATASET_CAMERA].to(device)
                    in_map: Tensor = batch[DATASET_MAP].to(device)
                    gt_label: Tensor = batch[DATASET_LABEL].to(device)
                    gt_depth: Tensor = batch[DATASET_DEPTH].to(device)

                    pred: PMOD_OUTPUT = pmodnet(
                        camera=in_camera, map_depth=in_map)
                    count_steps += in_camera.shape[0]

                    ae: METRIC_SUM_OUTPUT = metricSumAE(pred.depth, gt_depth)
                    sum_ae += ae.sum.sum().item()

                    se: METRIC_SUM_OUTPUT = metricSumSE(pred.depth, gt_depth)
                    sum_se += se.sum.sum().item()

                    ape: METRIC_SUM_OUTPUT = metricSumAPE(pred.depth, gt_depth)
                    sum_ape += ape.sum.sum().item()
                    count_in_range += ape.count.sum().item()

                    intersection: Tensor = metricSumIntersection(
                        pred.segmentation, gt_label)
                    sum_intersection += intersection.cpu().detach().numpy()

                    union: Tensor = metricSumUnion(pred.segmentation, gt_label)
                    sum_union += union.cpu().detach().numpy()

            tmp_mae: float = sum_ae / count_in_range * args[ARG_RANGE_NORM]
            tmp_rmse: float = np.sqrt(
                sum_se / count_in_range) * args[ARG_RANGE_NORM]
            tmp_mape: float = sum_ape / count_in_range
            tmp_miou: float = np.mean(sum_intersection / sum_union)

            val_metrics: Dict[str, float] = {}
            val_metrics['mIoU'] = tmp_miou
            val_metrics['MAE'] = tmp_mae
            val_metrics['RMSE'] = tmp_rmse
            val_metrics['MAPE'] = tmp_mape

            ###############
            # Show Result #
            ###############
            print(
                f'Epoch {epoch_num:5d}: {"Metric":14s}: {"mIoU":14s}: {tmp_miou * 100.0:8.2f} [ % ]')
            print(f'{"":11s}: {"":14s}: {"MAE":14s}: {tmp_mae:8.4f} [ m ]')
            print(f'{"":11s}: {"":14s}: {"RMSE":14s}: {tmp_rmse:8.4f} [ m ]')
            print(f'{"":11s}: {"":14s}: {"MAPE":14s}: {tmp_mape:8.4f}')

            ######################
            # Update TensorBoard #
            ######################
            if tb_writer is None:
                os.makedirs(tb_log_dir, exist_ok=True)
                tb_writer = tensorboard.SummaryWriter(log_dir=tb_log_dir)
            for h2, value in val_metrics.items():
                tb_writer.add_scalar(f'Validation/{h2}', value, epoch_num)
            tb_writer.flush()

            ###################
            # Save CheckPoint #
            ###################
            if os.path.isdir(checkpoint_dir) is False:
                os.makedirs(checkpoint_dir, exist_ok=True)
                with codecs.open(os.path.join(checkpoint_dir, 'config.yaml'), mode='w', encoding='utf-8') as f:
                    yaml.dump(args, f, encoding='utf-8', allow_unicode=True)
                workbook = xl.Workbook()
                worksheet = workbook.worksheets[0]
                worksheet.title = 'validation'
                worksheet.cell(row=1, column=1, value='Epoch')
                for col, title in enumerate(val_metrics.keys(), 2):
                    worksheet.cell(row=1, column=col, value=title)
                workbook.save(val_book_path)

            workbook = xl.load_workbook(val_book_path)
            worksheet = workbook.worksheets[0]
            worksheet.cell(row=epoch_itr+2, column=1, value=epoch_itr+1)
            for col, value in enumerate(val_metrics.values(), 2):
                worksheet.cell(row=epoch_itr+2, column=col, value=value)
            workbook.save(val_book_path)

            epoch_str: str = f'{epoch_num:05d}'
            if (best_miou.value < tmp_miou or best_mape.value > tmp_mape):
                pmodnet_ckpt_path: str = os.path.join(
                    checkpoint_dir, f'{epoch_str}_PMOD.pth')
                torch.save(pmodnet, pmodnet_ckpt_path)

                ###################
                # Auto Evaluation #
                ###################
                if args.get(ARG_AUTO_EVALUATION) is True:
                    eval_config_dict: Dict[str, Union[str, int]] = {
                        ARG_TAG: f'auto-eval-epoch{epoch_str}',
                        ARG_EVAL_DL_CONFIG: args[ARG_EVAL_DL_CONFIG],
                        ARG_BLOCK_SIZE: args[ARG_BLOCK_SIZE],
                        ARG_EVAL_DATA: args[ARG_EVAL_DATA],
                        ARG_NOMAP: args[ARG_NOMAP],
                        ARG_BATCH_SIZE: args[ARG_BATCH_SIZE],
                        ARG_CHECK_POINT: pmodnet_ckpt_path,
                        ARG_THRESHOLDS: args[ARG_THRESHOLDS],
                        ARG_TR_ERROR_RANGE: args[ARG_TR_ERROR_RANGE],
                        ARG_ROT_ERROR_RANGE: args[ARG_ROT_ERROR_RANGE]
                    }
                    eval_result_dict: Dict[str, Union[str, float]] = evaluate(
                        eval_config_dict, workdir, pmodnet, args)

                    # Prune out low-precision data.
                    tmp_datas: set = {eval_best_miou.path, eval_best_mape.path}
                    if eval_best_miou.value < eval_result_dict[METRIC_IOU]:
                        eval_best_miou = METRIC_BEST(
                            eval_result_dict[METRIC_IOU], eval_result_dict[DIR_RESULTS], epoch_num)
                    if eval_best_mape.value > eval_result_dict[METRIC_MAPE]:
                        eval_best_mape = METRIC_BEST(
                            eval_result_dict[METRIC_MAPE], eval_result_dict[DIR_RESULTS], epoch_num)
                    rm_datas: set = tmp_datas - \
                        {eval_best_miou.path, eval_best_mape.path}
                    for rm_data in rm_datas:
                        if os.path.isdir(rm_data):
                            shutil.rmtree(rm_data)

            # Remove "data.hdf5" for low-precision models.
            tmp_ckpts_epoch: set = {best_miou.epoch, best_mape.epoch}
            if best_miou.value < tmp_miou:
                best_miou = METRIC_BEST(tmp_miou, epoch=epoch_num)
            if best_mape.value > tmp_mape:
                best_mape = METRIC_BEST(tmp_mape, epoch=epoch_num)
            rm_ckpts_epoch: set = tmp_ckpts_epoch - \
                {best_miou.epoch, best_mape.epoch}
            eval_best_epoch: set = {eval_best_miou.epoch, eval_best_mape.epoch}
            for rm_ckpt_epoch in rm_ckpts_epoch:
                if {rm_ckpt_epoch} <= eval_best_epoch:
                    continue
                for ckpt_path in glob(os.path.join(checkpoint_dir, f'{rm_ckpt_epoch:05d}_PMOD_*.pth')):
                    if os.path.isfile(ckpt_path):
                        os.remove(ckpt_path)

            if isinstance(trial, optuna.Trial):
                if len(trial.study.directions) == 1:
                    trial.report(1.0 - tmp_miou + tmp_mape, epoch_itr)

            ########################
            # Update Learning Rate #
            ########################
            lr_dec_depth: float = update_lr(
                scheduler_dec_depth, optimizer_dec_depth, tmp_mape)
            lr_dec_seg: float = update_lr(
                scheduler_dec_seg, optimizer_dec_seg, 1.0 - tmp_miou)
            lr_enc: float = update_lr(
                scheduler_enc, optimizer_enc, (1.0 - tmp_miou) + tmp_mape)

            print(f'Epoch {epoch_num:5d}: {"Learning Rate":14s}: {"Generator":14s}: {"Encoder":14s}: {lr_enc:.4e}')
            print(f'{"":11s}: {"":14s}: {"":14s}: {"Decoder-Depth":14s}: {lr_dec_depth:.4e}')
            print(f'{"":11s}: {"":14s}: {"":14s}: {"Decoder-Seg":14s}: {lr_dec_seg:.4e}')

    except KeyboardInterrupt:
        print('KeyboardInterrupt')
    finally:
        if tb_writer is not None:
            tb_writer.close()

    return {METRIC_IOU: best_miou.value, METRIC_MAPE: best_mape.value}


if __name__ == "__main__":
    args = parse_args()
    main(args)
