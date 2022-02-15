import argparse
import codecs
import datetime
import os
from contextlib import redirect_stdout
import yaml
from tqdm import tqdm
import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet
import torch.nn as nn
from h5dataloader.common.structure import *
from h5datacreator import *
from .dataloader import *
from .model.metric import *


def parse_args() -> dict:
    parser = argparse.ArgumentParser()
    parser_eval = parser.add_argument_group('Evaluation')
    parser_eval.add_argument(
        '-t', f'--{arg_hyphen(ARG_TAG)}',
        type=str, required=True,
        help='Evaluation Tag.'
    )
    parser_eval.add_argument(
        '-cp', f'--{arg_hyphen(ARG_CHECK_POINT)}',
        type=str, metavar='PATH', required=True,
        help='PATH of checkpoint.'
    )
    parser_eval.add_argument(
        '-edc', f'--{arg_hyphen(ARG_EVAL_DL_CONFIG)}',
        type=str, metavar='PATH', required=True,
        help='PATH of JSON file of dataloader config.')
    parser_eval.add_argument(
        '-bs', f'--{arg_hyphen(ARG_BLOCK_SIZE)}',
        type=int, default=0,
        help='Block size of dataset.'
    )
    parser_eval.add_argument(
        '-ed', f'--{arg_hyphen(ARG_EVAL_DATA)}',
        type=str, metavar='PATH', nargs='+', required=True,
        help='PATH of evaluation HDF5 datasets.'
    )
    parser_eval.add_argument(
        f'--{arg_hyphen(ARG_NOMAP)}',
        action='store_true',
        help='No map input.'
    )
    parser_eval.add_argument(
        f'--{arg_hyphen(ARG_TRAIN_CONFIG)}',
        type=str, metavar='PATH', default=None,
        help='PATH of "config.yaml"'
    )
    parser_eval.add_argument(
        f'--{arg_hyphen(ARG_THRESHOLDS)}',
        type=float, nargs='+', default=DEFAULT_THRESHOLDS,
        help='Thresholds of depth.'
    )
    parser_eval.add_argument(
        f'--{arg_hyphen(ARG_SEED)}',
        type=int, default=0,
        help='Random seed.'
    )
    parser_net = parser.add_argument_group('Network')
    parser_net.add_argument(
        '-b', f'--{arg_hyphen(ARG_BATCH_SIZE)}',
        type=int, default=1,
        help='Batch Size'
    )

    args: dict = vars(parser.parse_args())

    return args


def _excel_detail_output(idx: int, batch_size: int, metric_results: Dict[str, METRIC_RESULT], excel_sheet: Worksheet) -> int:
    itr: int = idx
    for i in range(batch_size):
        row: int = itr + 3
        col: int = 1
        excel_sheet.cell(row=row, column=col, value=itr)
        col += 1
        for metric_key, metric_result in metric_results.items():
            if metric_key == METRIC_IOU:
                value: float = metric_result.all_metric[i].item() * 100.0
            else:
                value: float = metric_result.all_metric[i].item()
            excel_sheet.cell(row=row, column=col, value=value)
            col += 1

            if metric_result.use_class is True:
                for j in range(metric_result.class_metric.shape[1]):
                    if metric_key == METRIC_IOU:
                        value: float = metric_result.class_metric[i, j].item(
                        ) * 100.0
                    else:
                        value: float = metric_result.class_metric[i, j].item()
                    excel_sheet.cell(row=row, column=col, value=value)
                    col += 1
        itr += 1
    return itr


def _excel_overview_output(metric_results: Dict[str, METRIC_RESULT], excel_sheet: Worksheet, range_norm: float, result_dict: Dict[str, float] = None):
    for row, (metric_key, metric_result) in enumerate(metric_results.items(), 2):
        excel_sheet.cell(row=row, column=1, value=metric_result.name)

        if metric_key == METRIC_IOU:
            mean_all: torch.Tensor = (
                metric_result.sum_class_metric / metric_result.counts_metric).mean()
            mean_all_value: float = mean_all.item() * 100.0
        else:
            mean_all: torch.Tensor = metric_result.sum_class_metric.sum() / \
                metric_result.counts_metric.sum()

            if metric_key == METRIC_RMSE:
                mean_all_value: float = torch.sqrt(
                    mean_all).item() * range_norm
            else:
                mean_all_value: float = mean_all.item()

        excel_sheet.cell(row=row, column=2, value=mean_all_value)
        if isinstance(result_dict, dict):
            result_dict[metric_key] = mean_all_value

        if metric_result.use_class is True:
            col: int = 3

            if metric_result.use_class is True:
                if metric_result.counts_metric.ndim == 2:
                    mean_class: torch.Tensor = metric_result.sum_class_metric / \
                        metric_result.counts_metric.sum(dim=0)
                else:
                    mean_class: torch.Tensor = metric_result.sum_class_metric / \
                        metric_result.counts_metric

                for i in range(mean_class.shape[0]):
                    if metric_key == METRIC_RMSE:
                        excel_sheet.cell(row=row, column=col, value=torch.sqrt(
                            mean_class[i]).item() * range_norm)
                    elif metric_key == METRIC_IOU:
                        excel_sheet.cell(row=row, column=col,
                                         value=mean_class[i].item() * 100.0)
                    else:
                        excel_sheet.cell(row=row, column=col,
                                         value=mean_class[i].item())
                    col += 1

            if metric_result.use_thresholds is True:
                if metric_result.counts_metric.ndim == 2:
                    mean_threshold: torch.Tensor = metric_result.sum_threshold_metric / \
                        metric_result.counts_metric.sum(dim=1)
                else:
                    mean_threshold: torch.Tensor = metric_result.sum_threshold_metric / \
                        metric_result.counts_metric

                for i in range(mean_threshold.shape[0]):
                    if metric_key == METRIC_RMSE:
                        excel_sheet.cell(row=row, column=col, value=torch.sqrt(
                            mean_threshold[i]).item() * range_norm)
                    else:
                        excel_sheet.cell(row=row, column=col,
                                         value=mean_threshold[i].item())
                    col += 1


def _hdf5_output(h5_saver: H5Dataset, batch: Dict[str, torch.Tensor], pred: PMOD_OUTPUT, eval_dataset: PMOD_Test_Dataset, train_args: Dict[str, bool]):
    label_tag: str = eval_dataset.minibatch[DATASET_LABEL][CONFIG_TAG_LABELTAG]
    camera_norm_range_max: float = eval_dataset.minibatch[DATASET_CAMERA][CONFIG_TAG_RANGE][1] \
        if eval_dataset.minibatch[DATASET_CAMERA][CONFIG_TAG_NORMALIZE] is True else 1
    map_norm_range_max: float = eval_dataset.minibatch[DATASET_MAP][CONFIG_TAG_RANGE][1] \
        if eval_dataset.minibatch[DATASET_MAP][CONFIG_TAG_NORMALIZE] is True else 1.0
    depth_norm_range_max: float = eval_dataset.minibatch[DATASET_DEPTH][CONFIG_TAG_RANGE][1] \
        if eval_dataset.minibatch[DATASET_DEPTH][CONFIG_TAG_NORMALIZE] is True else 1.0

    for i in range(pred.depth.shape[0]):
        data_group = h5_saver.get_next_data_group()
        camera_img: np.ndarray = eval_dataset.to_numpy(
            batch[DATASET_CAMERA][i], DATASET_CAMERA) * camera_norm_range_max
        set_bgr8(data_group, 'Input-Camera',
                 np.uint8(np.where(camera_img > 255.0, 255.0, camera_img)),
                 eval_dataset.minibatch[DATASET_CAMERA][CONFIG_TAG_FRAMEID]
                 )
        set_depth(data_group, 'Input-Map',
                  eval_dataset.to_numpy(
                      batch[DATASET_MAP][i], DATASET_MAP) * map_norm_range_max,
                  eval_dataset.minibatch[DATASET_MAP][CONFIG_TAG_FRAMEID]
                  )
        set_semantic2d(data_group, 'GT-Label',
                       np.uint8(batch[DATASET_LABEL][i].numpy()),
                       eval_dataset.minibatch[DATASET_LABEL][CONFIG_TAG_FRAMEID],
                       label_tag
                       )
        set_depth(data_group, 'GT-Depth',
                  eval_dataset.to_numpy(
                      batch[DATASET_DEPTH][i], DATASET_DEPTH) * depth_norm_range_max,
                  eval_dataset.minibatch[DATASET_DEPTH][CONFIG_TAG_FRAMEID]
                  )
        set_semantic2d(data_group, 'Pred-Label',
                       np.uint8(pred.segmentation[i].argmax(
                           dim=0).cpu().detach().numpy()),
                       eval_dataset.minibatch[DATASET_LABEL][CONFIG_TAG_FRAMEID],
                       label_tag
                       )
        set_depth(data_group, 'Pred-Depth',
                  eval_dataset.to_numpy(
                      pred.depth[i], DATASET_DEPTH) * depth_norm_range_max,
                  eval_dataset.minibatch[DATASET_DEPTH][CONFIG_TAG_FRAMEID]
                  )


def main(args: Dict[str, Union[int, float, str, dict]], workdir: str, pmodnet: nn.Module = None, train_args: Dict[str, Union[int, float, str, dict]] = None):
    if os.path.isdir(workdir) is False:
        raise NotADirectoryError(f'{workdir}')

    ##################
    # Device Setting #
    ##################
    if torch.cuda.is_available():
        device: torch.device = torch.device('cuda')
        torch.backends.cudnn.benchmark = True
        print(f'{"Device":11s}: {torch.cuda.get_device_name(device)}')
    else:
        device: torch.device = torch.device('cpu')
        print(f'{"Device":11s}: CPU')

    ######################
    # DataLoader Setting #
    ######################
    if args[ARG_BLOCK_SIZE] < 3:
        args[ARG_BLOCK_SIZE] = 0

    with redirect_stdout(open(os.devnull, 'w')):
        eval_dataloader: DataLoader = init_eval_dataloader(args)
    eval_dataset: PMOD_Test_Dataset = eval_dataloader.dataset

    ###################
    # Network Setting #
    ###################
    if isinstance(pmodnet, nn.Module) is False:
        if os.path.isfile(args[ARG_CHECK_POINT]) is False:
            raise FileNotFoundError(f'{args[ARG_CHECK_POINT]}')
        pmodnet: nn.Module = torch.load(args[ARG_CHECK_POINT])
        pmodnet.to(device)

        if isinstance(args[ARG_TRAIN_CONFIG], str) is True:
            if os.path.isfile(args[ARG_TRAIN_CONFIG]) is False:
                raise FileNotFoundError(f'{args[ARG_TRAIN_CONFIG]}')
            config_path: str = args[ARG_TRAIN_CONFIG]
        else:
            config_path: str = os.path.join(
                os.path.dirname(args[ARG_CHECK_POINT]), 'config.yaml')
        with open(config_path) as f:
            train_args: dict = yaml.safe_load(f)
    else:
        if isinstance(train_args, dict) is False:
            raise ValueError(f'"train_args" must be dict.')
    pmodnet.eval()

    ##################
    # Metric Setting #
    ##################
    num_classes: int = args[ARG_NUM_CLASSES]
    range_norm: float = args[ARG_RANGE_NORM]
    thresholds: List[float] = args[ARG_THRESHOLDS]
    normalized_thresholds: List[float] = [th / range_norm for th in thresholds]

    metricIntersection: nn.Module = MetricIntersection().to(device)
    metricUnion: nn.Module = MetricUnion().to(device)
    metricBatchSumAE: nn.Module = MetricBatchSumAE(
        thresholds=normalized_thresholds, num_classes=num_classes).to(device)
    metricBatchSumSE: nn.Module = MetricBatchSumSE(
        thresholds=normalized_thresholds, num_classes=num_classes).to(device)
    metricBatchSumAPE: nn.Module = MetricBatchSumAPE(
        thresholds=normalized_thresholds, num_classes=num_classes).to(device)

    metric_results: Dict[str, METRIC_RESULT] = {
        METRIC_IOU: METRIC_RESULT(name='IoU [%]', use_class=True, all_tag='Mean'),
        METRIC_MAE: METRIC_RESULT(name='MAE [m]', use_class=True, use_thresholds=True),
        METRIC_RMSE: METRIC_RESULT(name='RMSE [m]', use_class=True, use_thresholds=True),
        METRIC_MAPE: METRIC_RESULT(name='MAPE', use_class=True, use_thresholds=True),
    }

    ###############
    # Save Config #
    ###############
    dt_start = datetime.datetime.now()
    args[ARG_DATE] = dt_start.strftime('%Y%m%dT%H%M%S')
    train_name: str = f'{train_args[ARG_DATE]}-{train_args[ARG_TAG]}'
    eval_name: str = f'{args[ARG_DATE]}-{args[ARG_TAG]}'
    result_dir: str = os.path.join(workdir, DIR_RESULTS, train_name, eval_name)
    os.makedirs(result_dir, exist_ok=True)

    # Excel Book
    eval_book = xl.Workbook()
    eval_sheet_overview = eval_book.worksheets[0]

    eval_sheet_overview.title = 'Overview'
    eval_sheet_overview.cell(row=1, column=1, value='Metric')
    eval_sheet_overview.cell(row=1, column=2, value='All')
    for col, (label, tag) in enumerate(args[ARG_LABEL_TAGS].items(), 3):
        eval_sheet_overview.cell(row=1, column=col, value=f'{label}:{tag}')
    tmp_thresholds = [0.0] + args[ARG_THRESHOLDS]
    for col, (th_min, th_max) in enumerate(zip(tmp_thresholds[:-1], tmp_thresholds[1:]), col+1):
        eval_sheet_overview.cell(
            row=1, column=col, value=f'{th_min}~{th_max}[m]')

    eval_sheet_detail = eval_book.create_sheet('Detail')
    eval_sheet_detail.cell(row=1, column=1, value='Step')
    eval_sheet_detail.merge_cells(
        start_row=1, end_row=2, start_column=1, end_column=1)

    metric_col = 2
    for metric_result in metric_results.values():
        eval_sheet_detail.cell(row=1, column=metric_col,
                               value=metric_result.name)
        if metric_result.use_class is True:
            cols: int = len(args[ARG_LABEL_TAGS]) + 1
            eval_sheet_detail.merge_cells(
                start_row=1, end_row=1, start_column=metric_col, end_column=metric_col+cols-1)
            eval_sheet_detail.cell(
                row=2, column=metric_col, value=metric_result.all_tag)
            for col, (label, tag) in enumerate(args[ARG_LABEL_TAGS].items(), metric_col + 1):
                eval_sheet_detail.cell(
                    row=2, column=col, value=f'{label}:{tag}')
        else:
            cols: int = 1
            eval_sheet_detail.merge_cells(
                start_row=1, end_row=2, start_column=metric_col, end_column=metric_col)
        metric_col += cols

    # HDF5
    h5_saver = H5Dataset(path=os.path.join(result_dir, 'data.hdf5'))
    label_tag: str = eval_dataset.minibatch[DATASET_LABEL][CONFIG_TAG_LABELTAG]
    label_group = h5_saver.get_label_group(label_tag)
    for label_config in eval_dataset.label_color_configs[label_tag]:
        bgr: np.ndarray = label_config[CONFIG_TAG_COLOR]
        set_label_config(
            label_group,
            label_config[CONFIG_TAG_LABEL],
            label_config[CONFIG_TAG_TAG],
            bgr[2], bgr[1], bgr[0]
        )

    # YAML
    with codecs.open(os.path.join(result_dir, 'config.yaml'), mode='w', encoding='utf-8') as f:
        yaml.dump(args, f, encoding='utf-8', allow_unicode=True)

    result_dict: Dict[str, Union[float, str]] = {DIR_RESULTS: result_dir}

    try:
        with torch.no_grad():
            ##################
            # Evluation Loop #
            ##################
            batch: Dict[str, torch.Tensor]

            itr: int = 0
            for batch in tqdm(eval_dataloader, desc=f'{"Evaluation":11s}'):
                if batch is None:
                    continue

                in_camera: torch.Tensor = batch[DATASET_CAMERA].to(device)
                gt_label: torch.Tensor = batch[DATASET_LABEL].to(device)
                gt_depth: torch.Tensor = batch[DATASET_DEPTH].to(device)
                if args[ARG_NOMAP] is True:
                    in_map: torch.Tensor = torch.full_like(batch[DATASET_MAP], fill_value=2.0).to(device)
                else:
                    in_map: torch.Tensor = batch[DATASET_MAP].to(device)

                pred: PMOD_OUTPUT = pmodnet(camera=in_camera, map_depth=in_map)

                # AE/MAE
                ae: METRIC_SUM_OUTPUT = metricBatchSumAE(
                    pred.depth, gt_depth, gt_label)
                metric_results[METRIC_MAE].class_metric = ae.sum.sum(
                    dim=1) / ae.count.sum(dim=1) * range_norm
                metric_results[METRIC_MAE].threshold_metric = ae.sum.sum(
                    dim=2) / ae.count.sum(dim=2) * range_norm
                metric_results[METRIC_MAE].all_metric = ae.sum.sum(
                    dim=(1, 2)) / ae.count.sum(dim=(1, 2)) * range_norm
                metric_results[METRIC_MAE].add(ae.count.sum(dim=0), ae.sum.sum(
                    dim=(0, 1)) * range_norm, ae.sum.sum(dim=(0, 2)) * range_norm)

                # SE/RMSE
                se: METRIC_SUM_OUTPUT = metricBatchSumSE(
                    pred.depth, gt_depth, gt_label)
                metric_results[METRIC_RMSE].class_metric = torch.sqrt(
                    se.sum.sum(dim=1) / se.count.sum(dim=1)) * range_norm
                metric_results[METRIC_RMSE].threshold_metric = torch.sqrt(
                    se.sum.sum(dim=2) / se.count.sum(dim=2)) * range_norm
                metric_results[METRIC_RMSE].all_metric = torch.sqrt(
                    se.sum.sum(dim=(1, 2)) / se.count.sum(dim=(1, 2))) * range_norm
                metric_results[METRIC_RMSE].add(se.count.sum(
                    dim=0), se.sum.sum(dim=(0, 1)), se.sum.sum(dim=(0, 2)))

                # APE/MAPE
                ape: METRIC_SUM_OUTPUT = metricBatchSumAPE(
                    pred.depth, gt_depth, gt_label)  # NTC
                metric_results[METRIC_MAPE].class_metric = ape.sum.sum(
                    dim=1) / ape.count.sum(dim=1)
                metric_results[METRIC_MAPE].threshold_metric = ape.sum.sum(
                    dim=2) / ape.count.sum(dim=2)
                metric_results[METRIC_MAPE].all_metric = ape.sum.sum(
                    dim=(1, 2)) / ape.count.sum(dim=(1, 2))
                metric_results[METRIC_MAPE].add(ape.count.sum(
                    dim=0), ape.sum.sum(dim=(0, 1)), ape.sum.sum(dim=(0, 2)))

                # IoU
                intersection: torch.Tensor = metricIntersection(
                    pred.segmentation, gt_label)
                union: torch.Tensor = metricUnion(pred.segmentation, gt_label)
                metric_results[METRIC_IOU].class_metric = intersection / union
                metric_results[METRIC_IOU].all_metric = metric_results[METRIC_IOU].class_metric.mean(
                    dim=1)
                metric_results[METRIC_IOU].add(
                    union.sum(dim=0), intersection.sum(dim=0))

                # -> Excel (Detail)
                itr = _excel_detail_output(
                    idx=itr, batch_size=in_camera.shape[0], metric_results=metric_results, excel_sheet=eval_sheet_detail)

                # -> HDF5
                _hdf5_output(h5_saver, batch, pred, eval_dataset, train_args)

            # -> Excel (Overview)
            _excel_overview_output(
                metric_results, eval_sheet_overview, range_norm, result_dict)
            print(
                f'{"":11s}: {"Metric":14s}: {"mIoU":14s}:{result_dict[METRIC_IOU]:9.2f} [ % ]')
            print(
                f'{"":11s}: {"":14s}: {"MAE":14s}:{result_dict[METRIC_MAE]:9.4f} [ m ]')
            print(
                f'{"":11s}: {"":14s}: {"RMSE":14s}:{result_dict[METRIC_RMSE]:9.4f} [ m ]')
            print(
                f'{"":11s}: {"":14s}: {"MAPE":14s}:{result_dict[METRIC_MAPE]:9.4f}')

    except KeyboardInterrupt:
        print('KeyboardInterrupt')
    finally:
        eval_book.save(os.path.join(result_dir, 'result.xlsx'))
        h5_saver.close()

    return result_dict
